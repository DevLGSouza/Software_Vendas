import openpyxl as op


exit = False
arquivo_bancodd = op.load_workbook("main\\banco_de_dados\\banco.xlsx")
banco = arquivo_bancodd["banco_de_dados"]
ultima_linha = banco.max_row


def vendashist():
    for linha in range(2,ultima_linha + 1):
        produtos = banco.cell(row=linha, column=1).value
        preços = banco.cell(row=linha, column=2).value
        datas = banco.cell(row=linha, column=3).value



    while True:
        print(f"{"\n"}{"\033[1;34mSALES HISTORY".center(42)}\n")
        print(f"{"DO YOU WANT TO SEARCH BY:  ".center(49)}{"\n"}")
        print("[1] PRODUCT NAME ".center(39))
        print("[2] PRODUCT PRICE ".center(39))
        print("[3] DATE OF SALE ".center(39))
        print("[4] RETURN TO MENU ".center(41))
        resp = input("\n")
        while True:
            if exit:
                break
            if resp == "1":
                resp2 = input(f"{"\n"}{"\033[1;36mENTER THE PRODUCT NAME: ".center(50)}{"\n"}{"\n"}\033[1;32m").strip().lower()
                if resp2 != "":
                    print(f"{"\n"}{"\033[1;36mPRODUCT HISTORY WITH NAME: ".center(47)}\033[1;32m{resp2}{"\n"}")
                    for linha in range(2,ultima_linha + 1):
                        produtos = banco.cell(row=linha, column=1).value
                        if resp2 == produtos:
                            print(f"{"PRODUCT: "}{produtos}{"  /  PRICE: R$"}{preços:.2f}{"  /  SALE DATE: "}{datas}")  
                    break                  
                else:
                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID NAME \033[m".center(51)}{"\n"}")

            elif resp == "2":
                resp3 = input(f"{"\n"}{"\033[1;36mENTER THE PRODUCT PRICE: ".center(50)}{"\n"}{"\n"}\033[1;32m").replace(",", ".").strip()
                try:
                    resp3float = float(resp3)
                    resp3str = str(resp3)
                    print(f"{"\n"}{"\033[1;36mPRODUCT HISTORY WITH PRICE: R$".center(45)}\033[1;32m{resp3float:.2f}{"\n"}")    
                    for linha in range(2,ultima_linha + 1):
                        preços = banco.cell(row=linha, column=2).value
                        if resp3 == preços:
                            print(f"{"PRODUCT: "}{produtos}{"  /  PRICE: R$"}{preços:.2f}{"  /  SALE DATE: "}{datas}")  
                    break    
                except ValueError:
                    print(f"{"\n"}{"\033[1;31m ERROR! ENTER A VALID PRICE \033[m".center(52)}{"\n"}")
                

            elif resp == "3":
                print(f"{"\n"}{"\033[1;36mENTER THE PRODUCT DATE: ".center(50)}{"\n"}\033[1;32m")
                resp4 = input("\033[1;36mFOR EXAMPLE: DD/MM/YYYY (DAY/MONTH/YEAR)\n\n".center(53)).replace("-", "/").replace(" ","/").replace(".","/").strip()
                print(f"{"\n"}{"\033[1;36mPRODUCT HISTORY WITH DATE: ".center(47)}\033[1;32m{resp4}{"\n"}")   
                print(f"{"\n"}{"\033[1;36mNAME , PRICE , DATE\033[m\n\033[1;32m"}")  
                for item in vendas:
                    if resp4 in item.strip().split(","):
                        print(item)
                break
            elif resp == "4":
                break
            else:
                print(f"{"\n"}{"\033[1;31m I CAN'T UNDERSTAND WHAT YOU WANT! ".center(49)}")
                print(f"\033[1;31mCHOOSE THE OPTION USING THE NUMBERS [1,2,3,4]".center(47))


vendashist()
